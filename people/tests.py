import io
import xml.etree.ElementTree as ET
from datetime import date
from urllib.parse import parse_qs

from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import IntegrityError, transaction
from django.test import TestCase
from django.urls import reverse

import openpyxl

from clubs.models import Club
from groups.models import Group, GroupMembership

from . import services
from .models import GuardianRelation, Membership, Person, StaffProfile
from .personnummer import _luhn, birth_date_from_personnummer
from .sportadmin import (
    import_person_rows,
    parse_personregister_xml,
    parse_sportadmin_personregister,
)


def complete_pnr(nine_digits):
    for digit in range(10):
        candidate = f"{nine_digits}{digit}"
        if _luhn(candidate):
            return candidate
    raise AssertionError("No valid checksum digit found")


class PersonnummerTests(TestCase):
    def setUp(self):
        self.club = Club.objects.create(name="Test BK", slug="testbk")

    def make_person(self, personnummer, **kwargs):
        defaults = dict(
            club=self.club,
            first_name="Anna",
            last_name="Andersson",
            gender=Person.Gender.FEMALE,
            street_address="Gatan 1",
            postal_code="11122",
            city="Stockholm",
        )
        defaults.update(kwargs)
        return Person.objects.create(personnummer=personnummer, **defaults)

    def test_save_normalizes_to_12_digits(self):
        pnr10 = complete_pnr("870512123")
        person = self.make_person(f"{pnr10[:6]}-{pnr10[6:]}")
        self.assertEqual(person.personnummer, "19870512" + pnr10[6:])
        self.assertEqual(person.birth_date, date(1987, 5, 12))

    def test_invalid_checksum_rejected(self):
        with self.assertRaises(ValidationError):
            self.make_person(complete_pnr("870512123")[:-1] + ("0" if complete_pnr("870512123")[-1] != "0" else "1"))

    def test_samordningsnummer_accepted(self):
        pnr10 = complete_pnr("870562123")
        person = self.make_person(pnr10)
        self.assertEqual(person.personnummer, "19870562" + pnr10[6:])
        self.assertEqual(person.birth_date, date(1987, 5, 2))

    def test_is_minor(self):
        adult = self.make_person(complete_pnr("870512123"), last_name="Adult")
        child = self.make_person(complete_pnr("160115123"), last_name="Child")
        self.assertFalse(adult.is_minor)
        self.assertTrue(child.is_minor)

    def test_masked_personnummer_accepted(self):
        person = self.make_person("20041003-****", last_name="Masked")
        self.assertEqual(person.personnummer, "20041003****")
        self.assertEqual(person.birth_date, date(2004, 10, 3))
        self.assertFalse(person.has_full_personnummer)

    def test_partial_personnummer_without_tail_accepted(self):
        person = self.make_person("20041003", last_name="Partial")
        self.assertEqual(person.personnummer, "20041003****")

    def test_masked_duplicates_allowed_per_club(self):
        first = self.make_person("20041003-****", last_name="Sibling A")
        second = self.make_person("20041003-****", last_name="Sibling B")
        self.assertEqual(first.personnummer, second.personnummer)

    def test_full_duplicate_still_rejected(self):
        # The uniqueness of *full* personnummer is enforced by a conditional
        # DB constraint (masked values are exempt), so this surfaces as an
        # IntegrityError on save rather than a ValidationError.
        pnr = complete_pnr("870512123")
        self.make_person(pnr, last_name="First")
        with self.assertRaises(IntegrityError), transaction.atomic():
            self.make_person(pnr, last_name="Second")

    def test_blank_personnummer_allowed(self):
        person = Person.objects.create(
            club=self.club,
            first_name="No",
            last_name="Pnr",
            gender="",
            street_address="",
            postal_code="",
            city="",
        )
        self.assertIsNone(person.personnummer)
        self.assertIsNone(person.birth_date)
        self.assertFalse(person.is_minor)


class MemberNumberTests(TestCase):
    def setUp(self):
        self.club = Club.objects.create(name="Test BK", slug="testbk")
        self.other_club = Club.objects.create(name="Other BK", slug="otherbk")

    def make_person(self, club, **kwargs):
        defaults = dict(
            club=club,
            first_name="Anna",
            last_name="Andersson",
            gender=Person.Gender.FEMALE,
        )
        defaults.update(kwargs)
        return Person.objects.create(**defaults)

    def test_member_number_auto_assigned(self):
        person = self.make_person(self.club)
        self.assertEqual(person.member_number, "0001")

    def test_member_number_increments_and_fills_gaps(self):
        first = self.make_person(self.club)
        second = self.make_person(self.club)
        self.assertEqual(first.member_number, "0001")
        self.assertEqual(second.member_number, "0002")
        first.delete()
        third = self.make_person(self.club)
        self.assertEqual(third.member_number, "0001")

    def test_member_number_scoped_per_club(self):
        a = self.make_person(self.club)
        b = self.make_person(self.other_club)
        # Same sequence position in different clubs -> identical number,
        # which is fine since uniqueness is enforced per club.
        self.assertEqual(a.member_number, "0001")
        self.assertEqual(b.member_number, "0001")

    def test_explicit_member_number_kept(self):
        person = self.make_person(self.club, member_number="X42")
        self.assertEqual(person.member_number, "X42")


def _build_sportadmin_xlsx(data_rows):
    """Build an in-memory xlsx with the Sportadmin Personregister layout."""
    headers = [
        "Personnummer", "Kön", "Förnamn", "Efternamn", "c/o", "Adress",
        "Postnummer", "Stad", "Land", "Mobiltelefon", "Telefon hem",
        "Telefon jobb", "E-post", "Målsman 1", "Relation", "E-post",
        "Telefon", "Målsman 2", "Relation", "E-post", "Telefon", "Skapad",
        "Uppdaterad", "Grupprekommendation", "Övrigt", "MedlemsNr",
        "StartÅr", "Allergi",
    ]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in data_rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class SportadminImportTests(TestCase):
    def setUp(self):
        self.club = Club.objects.create(name="Test BK", slug="testbk")

    def test_parse_maps_core_fields(self):
        content = _build_sportadmin_xlsx(
            [["20041003-****", "Man", "Abdulhadi", "Rasho", "", "Gatan 1",
              "80321", "GÄVLE", "Sverige", "0737461137", "", "",
              "abdul@example.com", "", "", "", "", "", "", "", "",
              "2025-12-05 08:23:30", "2025-12-05 08:23:30", "", "", "",
              "2025-04-10", ""]]
        )
        rows, warnings = parse_sportadmin_personregister(content)
        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["first_name"], "Abdulhadi")
        self.assertEqual(row["last_name"], "Rasho")
        self.assertEqual(row["personnummer"], "20041003****")
        self.assertEqual(row["gender"], Person.Gender.MALE)
        self.assertEqual(row["email"], "abdul@example.com")
        self.assertEqual(row["start_date"], "2025-04-10")
        self.assertEqual(warnings, [])

    def test_parse_guardian_block(self):
        content = _build_sportadmin_xlsx(
            [["20041003-****", "Kvinna", "Kid", "Small", "", "", "", "", "",
              "", "", "", "kid@example.com", "Mamma Pappa", "Mamma",
              "mamma@example.com", "0701234567", "", "", "", "", "", "", "",
              "", "", "", ""]]
        )
        rows, _ = parse_sportadmin_personregister(content)
        guardians = rows[0]["guardians"]
        self.assertEqual(len(guardians), 1)
        self.assertEqual(guardians[0]["name"], "Mamma Pappa")
        self.assertEqual(guardians[0]["relation"], "Mamma")
        self.assertEqual(guardians[0]["email"], "mamma@example.com")

    def test_import_creates_person_guardian_and_membership(self):
        rows = [{
            "first_name": "Kid",
            "last_name": "Small",
            "personnummer": None,
            "gender": Person.Gender.MALE,
            "street_address": "Gatan 1",
            "postal_code": "11122",
            "city": "Stockholm",
            "email": "kid@example.com",
            "phone_mobile": "0701112233",
            "notes": "",
            "allergy": "",
            "start_date": "2025-04-10",
            "guardians": [{"name": "Mamma Pappa", "relation": "Mamma",
                           "email": "mamma@example.com", "phone": "0709998877"}],
        }]
        created, skipped = import_person_rows(self.club, rows)
        self.assertEqual((created, skipped), (1, 0))
        person = Person.objects.get(email="kid@example.com")
        self.assertTrue(person.member_number)
        relation = GuardianRelation.objects.get(child=person)
        self.assertEqual(relation.relation, "Mamma")
        guardian = relation.guardian
        self.assertEqual(guardian.email, "mamma@example.com")
        self.assertIsNone(guardian.personnummer)
        membership = Membership.objects.get(person=person)
        self.assertEqual(membership.start_date, date(2025, 4, 10))

    def test_reimport_skips_existing(self):
        person = Person.objects.create(
            club=self.club,
            first_name="Anna",
            last_name="Andersson",
            gender=Person.Gender.FEMALE,
        )
        pnr = complete_pnr("850101123")
        person.personnummer = pnr
        person.save()
        rows = [{
            "first_name": "Anna",
            "last_name": "Andersson",
            "personnummer": pnr,
            "gender": "",
            "street_address": "",
            "postal_code": "",
            "city": "",
            "email": "",
            "phone_mobile": "",
            "notes": "",
            "allergy": "",
            "start_date": "",
            "guardians": [],
        }]
        created, skipped = import_person_rows(self.club, rows)
        self.assertEqual((created, skipped), (0, 1))


def _build_clubhub_xlsx(data_rows):
    """Build an in-memory xlsx with the ClubHub export layout."""
    headers = [
        "MedlemsNr", "Förnamn", "Efternamn", "Personnummer", "Kön",
        "Adress", "Postnummer", "Stad", "E-post", "Mobiltelefon",
        "Allergi", "Övrigt", "Roll", "Betalningsstatus", "Målsman",
        "Vårdnadshavare till",
    ]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Personregister"
    sheet.append(headers)
    for row in data_rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class ClubHubImportTests(TestCase):
    def setUp(self):
        self.club = Club.objects.create(name="Roundtrip BK", slug="roundtripbk")

    def test_parse_clubhub_xlsx(self):
        content = _build_clubhub_xlsx([[
            "0007", "Kid", "Small", "20041003****", "Kvinna", "Gatan 1",
            "80321", "Gävle", "kid@example.com", "0737461137", "Päron",
            "anteckning", "Medlem, Förälder", "Delvis betald",
            "Mamma Pappa (Mamma); Pappa Papa", "Small Kid",
        ]])
        rows, warnings = parse_sportadmin_personregister(content)
        self.assertEqual(warnings, [])
        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["first_name"], "Kid")
        self.assertEqual(row["personnummer"], "20041003****")
        self.assertEqual(row["gender"], Person.Gender.FEMALE)
        self.assertEqual(row["city"], "Gävle")
        self.assertEqual(row["allergy"], "Päron")
        self.assertEqual(row["notes"], "anteckning")
        self.assertEqual(row["roles"], ["Medlem", "Förälder"])
        self.assertEqual(row["payment_status"], "Delvis betald")
        self.assertEqual(
            row["guardians"],
            [
                {"name": "Mamma Pappa", "relation": "Mamma", "email": "", "phone": ""},
                {"name": "Pappa Papa", "relation": "", "email": "", "phone": ""},
            ],
        )
        self.assertEqual(row["children"], ["Small Kid"])

    def test_parse_clubhub_xml(self):
        xml = """<?xml version='1.0' encoding='UTF-8'?>
<personregister>
  <person>
    <fornamn>Kid</fornamn>
    <efternamn>Small</efternamn>
    <personnummer>20041003****</personnummer>
    <kon>Kvinna</kon>
    <epost>kid@example.com</epost>
    <roll>Medlem</roll>
    <betalningsstatus>Betald</betalningsstatus>
    <malsman>Mamma Pappa (Mamma)</malsman>
  </person>
  <person>
    <fornamn>Mamma</fornamn>
    <efternamn>Pappa</efternamn>
    <epost>mamma@example.com</epost>
    <roll>Förälder</roll>
    <vardnadshavare>Kid Small</vardnadshavare>
  </person>
</personregister>"""
        rows, warnings = parse_personregister_xml(xml)
        self.assertEqual(warnings, [])
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["roles"], ["Medlem"])
        self.assertEqual(rows[0]["payment_status"], "Betald")
        self.assertEqual(rows[0]["guardians"][0]["name"], "Mamma Pappa")
        self.assertEqual(rows[0]["guardians"][0]["relation"], "Mamma")
        self.assertEqual(rows[1]["roles"], ["Förälder"])
        self.assertEqual(rows[1]["children"], ["Kid Small"])

        created, skipped = import_person_rows(self.club, rows)
        self.assertEqual((created, skipped), (1, 1))
        kid = Person.objects.get(first_name="Kid")
        self.assertTrue(Membership.objects.filter(person=kid).exists())
        self.assertEqual(
            kid.membership.payment_status, Membership.PaymentStatus.PAID
        )
        relation = GuardianRelation.objects.get(child=kid)
        self.assertEqual(relation.relation, "Mamma")
        self.assertEqual(relation.guardian.email, "mamma@example.com")

    def test_import_clubhub_roles(self):
        rows = [{
            "first_name": "Erik",
            "last_name": "Expert",
            "personnummer": None,
            "gender": "",
            "street_address": "",
            "postal_code": "",
            "city": "",
            "email": "erik@example.com",
            "phone_mobile": "",
            "notes": "",
            "allergy": "",
            "start_date": "",
            "payment_status": "Obetald",
            "roles": ["Medlem", "Tränare", "Admin"],
            "guardians": [],
            "children": ["Kid Small"],
        }]
        created, skipped = import_person_rows(self.club, rows)
        self.assertEqual((created, skipped), (1, 0))
        person = Person.objects.get(email="erik@example.com")
        membership = Membership.objects.get(person=person)
        self.assertEqual(membership.payment_status, Membership.PaymentStatus.UNPAID)
        profile = StaffProfile.objects.get(person=person)
        self.assertTrue(profile.is_admin)
        child = Person.objects.get(first_name="Kid")
        self.assertTrue(
            GuardianRelation.objects.filter(guardian=person, child=child).exists()
        )

    def test_import_without_member_role_creates_no_membership(self):
        rows = [{
            "first_name": "Bara",
            "last_name": "Förälder",
            "personnummer": None,
            "gender": "",
            "street_address": "",
            "postal_code": "",
            "city": "",
            "email": "",
            "phone_mobile": "",
            "notes": "",
            "allergy": "",
            "start_date": "",
            "payment_status": "",
            "roles": ["Förälder"],
            "guardians": [],
            "children": [],
        }]
        import_person_rows(self.club, rows)
        person = Person.objects.get(first_name="Bara")
        self.assertFalse(Membership.objects.filter(person=person).exists())
        self.assertFalse(StaffProfile.objects.filter(person=person).exists())


class MembershipGuardianRuleTests(TestCase):
    def setUp(self):
        self.club = Club.objects.create(name="Test BK", slug="testbk")

    def make_person(self, personnummer, first_name, last_name):
        return Person.objects.create(
            club=self.club,
            first_name=first_name,
            last_name=last_name,
            personnummer=personnummer,
            gender=Person.Gender.FEMALE,
            street_address="Gatan 1",
            postal_code="11122",
            city="Stockholm",
        )

    def test_minor_requires_guardian(self):
        adult = self.make_person(complete_pnr("800101123"), "Guard", "Ian")
        child = self.make_person(complete_pnr("160115123"), "Kid", "Small")
        membership = Membership(person=child, start_date=date(2026, 1, 1))
        with self.assertRaises(ValidationError):
            membership.full_clean()
        GuardianRelation.objects.create(guardian=adult, child=child)
        membership.full_clean()

    def test_adult_needs_no_guardian(self):
        adult = self.make_person(complete_pnr("800101123"), "Big", "Person")
        Membership(person=adult, start_date=date(2026, 1, 1)).full_clean()


class AccessServiceTests(TestCase):
    def setUp(self):
        self.club = Club.objects.create(name="Test BK", slug="testbk")

    def make_user_person(self, username, is_admin=False, nine_digits="850101123"):
        user = User.objects.create_user(username=username, password="pw12345!")
        person = Person.objects.create(
            club=self.club,
            first_name="A",
            last_name=username.title(),
            personnummer=complete_pnr(nine_digits),
            gender=Person.Gender.FEMALE,
            street_address="Gatan 1",
            postal_code="11122",
            city="Stockholm",
            user=user,
        )
        StaffProfile.objects.create(person=person, is_admin=is_admin)
        return user, person

    def test_plain_user_has_no_person(self):
        user = User.objects.create_user(username="plain", password="pw12345!")
        self.assertIsNone(services.get_person(user))
        self.assertFalse(services.has_staff_profile(user))
        self.assertFalse(services.is_admin(user))

    def test_is_admin_flag_respected(self):
        admin_user, _ = self.make_user_person("boss", is_admin=True, nine_digits="820202123")
        trainer_user, _ = self.make_user_person("coach", is_admin=False, nine_digits="870512123")
        self.assertTrue(services.is_admin(admin_user))
        self.assertFalse(services.is_admin(trainer_user))
        self.assertTrue(services.has_staff_profile(trainer_user))

    def test_visible_groups_scoping(self):
        trainer_user, trainer_person = self.make_user_person("coach", nine_digits="870512123")
        admin_user, admin_person = self.make_user_person("boss", is_admin=True, nine_digits="820202123")
        own = Group.objects.create(club=self.club, name="Egna")
        other = Group.objects.create(club=self.club, name="Främmande")
        GroupMembership.objects.create(
            group=own, person=trainer_person, role=GroupMembership.Role.TRAINER
        )
        self.assertEqual(list(services.visible_groups(trainer_user)), [own])
        self.assertEqual(set(services.visible_groups(admin_user)), {own, other})
        self.assertTrue(services.can_manage_group(trainer_user, own))
        self.assertFalse(services.can_manage_group(trainer_user, other))


class AuthFlowTests(TestCase):
    def setUp(self):
        self.club = Club.objects.create(name="Login BK", slug="loginbk")
        self.user = User.objects.create_user(username="anna", password="starkpass1")
        person = Person.objects.create(
            club=self.club,
            first_name="Anna",
            last_name="Ekström",
            personnummer=complete_pnr("850101123"),
            gender=Person.Gender.FEMALE,
            street_address="Gatan 1",
            postal_code="11122",
            city="Stockholm",
            user=self.user,
        )
        StaffProfile.objects.create(person=person)

    def test_login_page_reachable(self):
        response = self.client.get(reverse("login"))
        self.assertEqual(response.status_code, 200)

    def test_login_redirects_to_dashboard(self):
        response = self.client.post(
            reverse("login"), {"username": "anna", "password": "starkpass1"}
        )
        self.assertRedirects(response, reverse("clubs:home"))
        page = self.client.get(reverse("clubs:home"))
        self.assertContains(page, "Hej Anna")

    def test_home_forbidden_without_profile(self):
        outsider = User.objects.create_user(username="out", password="starkpass1")
        self.client.force_login(outsider)
        self.assertEqual(self.client.get(reverse("clubs:home")).status_code, 403)

    def test_home_requires_login(self):
        response = self.client.get(reverse("clubs:home"))
        self.assertEqual(response.status_code, 302)


class PersonRegisterViewTests(TestCase):
    def setUp(self):
        self.club = Club.objects.create(name="Register BK", slug="registerbk")
        self.admin_user = User.objects.create_user(username="boss", password="pw12345!")
        admin_person = Person.objects.create(
            club=self.club,
            first_name="Anna",
            last_name="Admin",
            user=self.admin_user,
        )
        StaffProfile.objects.create(person=admin_person, is_admin=True)
        self.client.force_login(self.admin_user)

    def make_person(self, first_name, last_name, **kwargs):
        return Person.objects.create(
            club=self.club, first_name=first_name, last_name=last_name, **kwargs
        )

    @staticmethod
    def pnr_for_age(age):
        """Valid personnummer for a person born on Jan 1, ~age years ago."""
        year = (date.today().year - age) % 100
        return complete_pnr(f"{year:02d}0101123")

    def get(self, **params):
        return self.client.get(reverse("people:register"), params)

    def test_search_matches_name_and_member_number(self):
        bengt = self.make_person("Bengt", "Berg")
        self.make_person("Cecilia", "Carlsson")
        response = self.get(q="bengt")
        self.assertContains(response, "Berg")
        self.assertNotContains(response, "Carlsson")
        response = self.get(q=bengt.member_number)
        self.assertContains(response, "Bengt")
        self.assertNotContains(response, "Cecilia")

    def test_search_matches_full_name(self):
        self.make_person("Anna", "Svensson")
        self.make_person("Anna", "Andersson")
        response = self.get(q="Anna Svensson")
        self.assertContains(response, "Svensson")
        self.assertNotContains(response, "Andersson")

    def test_headers_are_sort_links_with_arrow_icons(self):
        response = self.get()
        self.assertContains(response, "th-sort-link", count=5)
        self.assertContains(response, "fa-arrow-down")

    def test_sort_column_toggles_direction_and_icon(self):
        columns = {col["key"]: col for col in self.get().context["columns"]}
        self.assertEqual(columns["name"]["icon"], "fa-arrow-down")
        self.assertIn("sort=name", columns["name"]["url"])
        self.assertIn("dir=desc", columns["name"]["url"])

        columns = {
            col["key"]: col
            for col in self.get(sort="name", dir="desc").context["columns"]
        }
        self.assertEqual(columns["name"]["icon"], "fa-arrow-up")
        self.assertIn("dir=asc", columns["name"]["url"])

    def test_sort_by_member_number_and_name(self):
        self.make_person("Cecilia", "Carlsson")
        self.make_person("Bengt", "Berg")
        response = self.get(sort="nr", dir="asc")
        names = [p.last_name for p in response.context["people"]]
        # The admin was created first -> member number 0001.
        self.assertEqual(names, ["Admin", "Carlsson", "Berg"])
        response = self.get(sort="name", dir="desc")
        names = [p.last_name for p in response.context["people"]]
        self.assertEqual(names, ["Carlsson", "Berg", "Admin"])

    def test_sort_preserves_search_and_filters(self):
        self.make_person("Bengt", "Berg", email="berg@example.com")
        response = self.get(q="berg", sort="email", dir="desc")
        self.assertIn("q=berg", response.context["columns"][0]["url"])
        self.assertIn("sort=nr", response.context["columns"][0]["url"])
        self.assertIn("dir=desc", response.context["columns"][0]["url"])

    def test_role_filters(self):
        member = self.make_person("Mimmi", "Medlem")
        Membership.objects.create(person=member, start_date=date(2026, 1, 1))
        trainer = self.make_person("Tore", "Tranare")
        StaffProfile.objects.create(person=trainer)
        group = Group.objects.create(club=self.club, name="Grupp A")
        GroupMembership.objects.create(
            group=group, person=trainer, role=GroupMembership.Role.TRAINER
        )
        child = self.make_person("Kalle", "Barn")
        guardian = self.make_person("Gunilla", "Foralder")
        GuardianRelation.objects.create(guardian=guardian, child=child)

        response = self.get(role="member")
        self.assertEqual({p.last_name for p in response.context["people"]}, {"Medlem"})
        response = self.get(role="trainer")
        self.assertEqual({p.last_name for p in response.context["people"]}, {"Tranare"})
        response = self.get(role="guardian")
        self.assertEqual({p.last_name for p in response.context["people"]}, {"Foralder"})
        response = self.get(role=["member", "trainer"])
        self.assertEqual(
            {p.last_name for p in response.context["people"]},
            {"Medlem", "Tranare"},
        )
        # "Other" = none of the role markers; the admin and the child
        # (guardianship marks the guardian, not the child) are plain persons.
        response = self.get(role="other")
        self.assertEqual({p.last_name for p in response.context["people"]}, {"Admin", "Barn"})
        # Every role selected, including "other", matches everyone.
        response = self.get(role=["member", "trainer", "guardian", "other"])
        self.assertEqual(
            {p.last_name for p in response.context["people"]},
            {"Medlem", "Tranare", "Foralder", "Admin", "Barn"},
        )

    def test_payment_filters(self):
        payer = self.make_person("Par", "Betald")
        Membership.objects.create(
            person=payer,
            start_date=date(2026, 1, 1),
            payment_status=Membership.PaymentStatus.PAID,
        )
        debtor = self.make_person("Doris", "Obetald")
        Membership.objects.create(
            person=debtor,
            start_date=date(2026, 1, 1),
            payment_status=Membership.PaymentStatus.UNPAID,
        )
        halffer = self.make_person("Hugo", "Halvbetald")
        Membership.objects.create(
            person=halffer,
            start_date=date(2026, 1, 1),
            payment_status=Membership.PaymentStatus.PARTLY_PAID,
        )
        response = self.get(payment="paid")
        self.assertEqual({p.last_name for p in response.context["people"]}, {"Betald"})
        response = self.get(payment="partly")
        self.assertEqual({p.last_name for p in response.context["people"]}, {"Halvbetald"})
        response = self.get(payment="unpaid")
        self.assertEqual({p.last_name for p in response.context["people"]}, {"Obetald"})
        response = self.get(payment=["paid", "partly", "unpaid"])
        self.assertEqual(
            {p.last_name for p in response.context["people"]},
            {"Betald", "Halvbetald", "Obetald"},
        )

    def test_gender_filters(self):
        man = self.make_person("Mats", "Man", gender=Person.Gender.MALE)
        woman = self.make_person("Karin", "Kvinna", gender=Person.Gender.FEMALE)
        non_binary = self.make_person("Kim", "Ickebinary", gender=Person.Gender.NON_BINARY)
        self.make_person("Nils", "Okant")  # no gender set

        response = self.get(gender="male")
        self.assertEqual({p.last_name for p in response.context["people"]}, {"Man"})
        response = self.get(gender="female")
        self.assertEqual({p.last_name for p in response.context["people"]}, {"Kvinna"})
        response = self.get(gender="non_binary")
        self.assertEqual({p.last_name for p in response.context["people"]}, {"Ickebinary"})
        response = self.get(gender=["male", "female", "non_binary"])
        self.assertEqual(
            {p.last_name for p in response.context["people"]},
            {"Man", "Kvinna", "Ickebinary"},
        )
        # All three unselected clears the gender filter entirely.
        response = self.get(gender="")
        self.assertEqual(
            {p.last_name for p in response.context["people"]},
            {man.last_name, woman.last_name, non_binary.last_name, "Okant", "Admin"},
        )

    def test_age_filters(self):
        self.make_person("Astrid", "Vuxen", personnummer=self.pnr_for_age(20))
        self.make_person("Kenny", "Barn", personnummer=self.pnr_for_age(3))
        # Without personnummer the age cannot be determined -> matches neither.
        self.make_person("Namnlos", "Okand")
        response = self.get(age="over15")
        self.assertEqual({p.last_name for p in response.context["people"]}, {"Vuxen"})
        response = self.get(age="under15")
        self.assertEqual({p.last_name for p in response.context["people"]}, {"Barn"})

    def test_unknown_sort_params_fall_back_to_member_number(self):
        response = self.get(sort="hacker", dir="sideways")
        self.assertEqual(response.context["sort"], "nr")
        self.assertEqual(response.context["dir"], "asc")

    def test_defaults_show_everyone_sorted_by_member_number(self):
        member = self.make_person(
            "Mimmi", "Medlem", personnummer=self.pnr_for_age(20),
            gender=Person.Gender.FEMALE,
        )
        Membership.objects.create(person=member, start_date=date(2026, 1, 1))
        guardian = self.make_person("Gunilla", "Foralder")
        child = self.make_person("Kalle", "Barn", personnummer=self.pnr_for_age(3))
        GuardianRelation.objects.create(guardian=guardian, child=child)
        untagged = self.make_person("Nils", "Ingen")

        # First visit: no query params at all -> no filters applied, so
        # members, guardians, children and untagged persons all show up.
        response = self.client.get(reverse("people:register"))
        self.assertEqual((response.context["sort"], response.context["dir"]), ("nr", "asc"))
        self.assertEqual(
            {p.last_name for p in response.context["people"]},
            {"Medlem", "Foralder", "Barn", "Ingen", "Admin"},
        )
        self.assertFalse(response.context["is_filtered"])
        self.assertEqual(
            response.context["active_filters"],
            {"role": [], "payment": [], "age": [], "gender": []},
        )

    def test_filtered_empty_shows_match_message(self):
        response = self.get(q="finnsinte")
        self.assertContains(response, "Inga personer matchar sökningen eller filtret.")
        self.assertNotContains(response, "Inga personer registrerade ännu.")

    def test_export_button_and_modal_render(self):
        response = self.get()
        self.assertContains(response, 'data-bs-target="#export-modal"')
        self.assertContains(response, 'data-bs-target="#export-modal"')
        self.assertContains(response, 'id="export-modal"')
        self.assertContains(response, 'value="xlsx" checked')
        self.assertContains(response, 'value="xml"')
        # The modal spells out that active search/filters carry over.
        self.assertContains(response, "aktiva just nu.")

    def test_pagination_defaults_and_page_size_links(self):
        response = self.get()
        self.assertEqual(response.context["per_page"], 40)
        options = {o["value"]: o for o in response.context["per_page_options"]}
        self.assertEqual(set(options), {40, 60, 100})
        self.assertTrue(options[40]["active"])
        self.assertFalse(options[60]["active"])
        query = parse_qs(options[60]["url"].lstrip("?"))
        self.assertEqual(query["per_page"], ["60"])
        # Changing page size resets to the first page.
        self.assertNotIn("page", query)

    def test_invalid_per_page_falls_back_to_default(self):
        response = self.get(per_page="9999")
        self.assertEqual(response.context["per_page"], 40)
        response = self.get(per_page="nonsense")
        self.assertEqual(response.context["per_page"], 40)

    def test_pagination_pages_and_entry_counts(self):
        for index in range(45):
            self.make_person("Först", f"Person{index:02d}")
        # 45 created + the admin = 46 people.
        response = self.get()
        self.assertEqual(response.context["entries_total"], 46)
        self.assertEqual(response.context["entries_start"], 1)
        self.assertEqual(response.context["entries_end"], 40)
        query = parse_qs(response.context["next_page_url"].lstrip("?"))
        self.assertEqual(query["page"], ["2"])
        self.assertIsNone(response.context["prev_page_url"])
        self.assertEqual(len(response.context["people"]), 40)

        response = self.get(page="2")
        self.assertEqual(response.context["entries_start"], 41)
        self.assertEqual(response.context["entries_end"], 46)
        self.assertEqual(len(response.context["people"]), 6)
        query = parse_qs(response.context["prev_page_url"].lstrip("?"))
        self.assertEqual(query["page"], ["1"])
        self.assertIsNone(response.context["next_page_url"])
        # Page-size links keep working from page 2 (and drop the page param).
        options = {o["value"]: o for o in response.context["per_page_options"]}
        query = parse_qs(options[100]["url"].lstrip("?"))
        self.assertEqual(query["per_page"], ["100"])
        self.assertNotIn("page", query)

    def test_pagination_preserves_search_and_sort(self):
        for index in range(45):
            self.make_person("Först", f"Person{index:02d}", email="berg@example.com")
        response = self.get(q="berg", sort="nr", dir="desc", page="1")
        # Search survives pagination links.
        query = parse_qs(response.context["next_page_url"].lstrip("?"))
        self.assertEqual(query.get("q"), ["berg"])
        self.assertEqual(query.get("sort"), ["nr"])
        self.assertEqual(query.get("dir"), ["desc"])
        # Sorting resets to page 1; the active column toggles back to asc.
        columns = {col["key"]: col for col in response.context["columns"]}
        query = parse_qs(columns["nr"]["url"].lstrip("?"))
        self.assertNotIn("page", query)
        self.assertEqual(query.get("q"), ["berg"])
        self.assertEqual(query.get("dir"), ["asc"])
        self.assertEqual(query.get("sort"), ["nr"])



class PersonExportViewTests(TestCase):
    def setUp(self):
        self.club = Club.objects.create(name="Export BK", slug="exportbk")
        self.admin_user = User.objects.create_user(username="boss", password="pw12345!")
        admin_person = Person.objects.create(
            club=self.club,
            first_name="Anna",
            last_name="Admin",
            email="anna@example.com",
            user=self.admin_user,
        )
        StaffProfile.objects.create(person=admin_person, is_admin=True)
        self.client.force_login(self.admin_user)
        self.url = reverse("people:export")

    def make_person(self, first_name, last_name, **kwargs):
        return Person.objects.create(
            club=self.club, first_name=first_name, last_name=last_name, **kwargs
        )

    def test_requires_admin(self):
        self.client.logout()
        self.assertEqual(self.client.get(self.url).status_code, 302)
        outsider = User.objects.create_user(username="out", password="pw12345!")
        self.client.force_login(outsider)
        self.assertEqual(self.client.get(self.url).status_code, 403)
        # Plain staff (no admin flag) is not enough either.
        staff_user = User.objects.create_user(username="staff", password="pw12345!")
        staff_person = Person.objects.create(
            club=self.club, first_name="Sten", last_name="Staff", user=staff_user
        )
        StaffProfile.objects.create(person=staff_person)
        self.client.force_login(staff_user)
        self.assertEqual(self.client.get(self.url).status_code, 403)

    def test_xlsx_export_contains_header_and_rows(self):
        berg_pnr = complete_pnr("041003123")
        self.make_person(
            "Bengt", "Berg",
            personnummer=berg_pnr,
            email="berg@example.com", phone_mobile="0701234567",
        )
        self.make_person("Cecilia", "Carlsson")
        response = self.client.get(self.url, {"format": "xlsx"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn("personregister-", response["Content-Disposition"])
        self.assertIn(".xlsx", response["Content-Disposition"])

        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        self.assertEqual(
            list(rows[0]),
            [
                "MedlemsNr", "Förnamn", "Efternamn", "Personnummer", "Kön",
                "Adress", "Postnummer", "Stad", "E-post", "Mobiltelefon",
                "Allergi", "Övrigt", "Roll", "Betalningsstatus", "Målsman",
                "Vårdnadshavare till",
            ],
        )
        # Header plus every person in the club, admin included, in
        # member-number order.
        self.assertEqual(len(rows), 4)
        self.assertEqual(rows[1][1:3], ("Anna", "Admin"))
        self.assertEqual(rows[2][1:3], ("Bengt", "Berg"))
        self.assertEqual(rows[2][3], "20" + berg_pnr)
        self.assertEqual(rows[2][8], "berg@example.com")
        self.assertEqual(rows[2][9], "0701234567")

    def test_xlsx_export_computed_columns_and_roundtrip(self):
        kid = self.make_person(
            "Kid", "Small", gender=Person.Gender.FEMALE, allergy="Päron"
        )
        guardian = self.make_person("Mamma", "Pappa", email="mamma@example.com")
        GuardianRelation.objects.create(
            guardian=guardian, child=kid, relation="Mamma"
        )
        Membership.objects.create(
            person=kid,
            start_date=date(2026, 1, 1),
            payment_status=Membership.PaymentStatus.PAID,
        )
        group = Group.objects.create(club=self.club, name="Tävlingsgrupp")
        GroupMembership.objects.create(
            person=kid, group=group, role=GroupMembership.Role.TRAINER
        )

        response = self.client.get(self.url, {"format": "xlsx"})
        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        by_name = {row[1] + " " + row[2]: row for row in rows[1:]}
        self.assertEqual(by_name["Kid Small"][12], "Medlem, Tränare")
        self.assertEqual(by_name["Kid Small"][13], "Betald")
        self.assertEqual(by_name["Kid Small"][14], "Mamma Pappa (Mamma)")
        self.assertEqual(by_name["Kid Small"][10], "Päron")
        self.assertEqual(by_name["Mamma Pappa"][12], "Förälder")
        self.assertEqual(by_name["Mamma Pappa"][15], "Kid Small")

        # The export re-imports cleanly into a fresh club. The guardian is
        # created from the child's "Målsman" text, so her own row matches it
        # by name and is merged (email backfilled) rather than duplicated.
        other = Club.objects.create(name="Other BK", slug="otherbk")
        parsed, _ = parse_sportadmin_personregister(response.content)
        created, skipped = import_person_rows(other, parsed)
        self.assertEqual((created, skipped), (2, 1))
        self.assertEqual(
            Person.objects.filter(club=other).count(), 3
        )
        new_kid = Person.objects.get(club=other, first_name="Kid")
        self.assertEqual(new_kid.allergy, "Päron")
        self.assertEqual(new_kid.gender, Person.Gender.FEMALE)
        self.assertEqual(new_kid.membership.payment_status, Membership.PaymentStatus.PAID)
        self.assertTrue(StaffProfile.objects.filter(person=new_kid).exists())
        relation = GuardianRelation.objects.get(child=new_kid)
        self.assertEqual(relation.relation, "Mamma")

    def test_xml_export_is_valid_and_nested(self):
        self.make_person("Bengt", "Berg", email="berg@example.com")
        response = self.client.get(self.url, {"format": "xml"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/xml; charset=utf-8")
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(".xml", response["Content-Disposition"])

        root = ET.fromstring(response.content)
        self.assertEqual(root.tag, "personregister")
        persons = root.findall("person")
        self.assertEqual(len(persons), 2)
        bengt = persons[1]
        self.assertEqual(bengt.findtext("fornamn"), "Bengt")
        self.assertEqual(bengt.findtext("efternamn"), "Berg")
        self.assertEqual(bengt.findtext("epost"), "berg@example.com")
        self.assertEqual(bengt.findtext("personnummer"), "")

    def test_export_respects_search_and_filters(self):
        member = self.make_person("Mimmi", "Medlem", email="mimmi@example.com")
        Membership.objects.create(person=member, start_date=date(2026, 1, 1))
        self.make_person("Gunilla", "Foralder")

        def exported_names(**params):
            response = self.client.get(self.url, {"format": "xml", **params})
            root = ET.fromstring(response.content)
            return {p.findtext("efternamn") for p in root.findall("person")}

        # No params: everything (default state is no filtering at all).
        self.assertEqual(exported_names(), {"Admin", "Medlem", "Foralder"})
        # Role filter narrows the export (the admin is not a member)...
        self.assertEqual(exported_names(role="member"), {"Medlem"})
        # ...as does the search box.
        self.assertEqual(exported_names(q="mimmi"), {"Medlem"})
        # Sort is honoured too.
        response = self.client.get(self.url, {"format": "xml", "sort": "name", "dir": "desc"})
        root = ET.fromstring(response.content)
        self.assertEqual(
            [p.findtext("efternamn") for p in root.findall("person")],
            ["Medlem", "Foralder", "Admin"],
        )

    def test_export_ignores_pagination(self):
        for index in range(45):
            self.make_person("Först", f"Person{index:02d}")
        response = self.client.get(self.url, {"format": "xml", "page": "2", "per_page": "40"})
        root = ET.fromstring(response.content)
        self.assertEqual(len(root.findall("person")), 46)

    def test_unknown_format_falls_back_to_xlsx(self):
        response = self.client.get(self.url, {"format": "exe"})
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class PersonImportPreviewTests(TestCase):
    """Selection checkboxes and the preview filter in the import modal."""

    def setUp(self):
        self.club = Club.objects.create(name="Import BK", slug="importbk")
        self.admin_user = User.objects.create_user(username="boss", password="pw12345!")
        admin_person = Person.objects.create(
            club=self.club,
            first_name="Anna",
            last_name="Admin",
            user=self.admin_user,
        )
        StaffProfile.objects.create(person=admin_person, is_admin=True)
        self.client.force_login(self.admin_user)

    def upload_sportadmin(self, data_rows):
        content = _build_sportadmin_xlsx(data_rows)
        response = self.client.post(
            reverse("people:import_preview"),
            {"file": SimpleUploadedFile("register.xlsx", content)},
        )
        self.assertEqual(response.status_code, 200)
        return response

    def confirm(self, **post):
        return self.client.post(reverse("people:import_confirm"), post, follow=True)

    def test_preview_renders_all_rows_checked(self):
        response = self.upload_sportadmin(
            [
                ["20041003-****", "Kvinna", "Adult", "Person", "", "", "", "", "",
                 "", "", "", "adult@example.com", "", "", "", "", "", "", "", "",
                 "", "", "", "", "", "2025-04-10", ""],
                ["20180101-****", "Man", "Kid", "Person", "", "", "", "", "",
                 "", "", "", "", "", "", "", "", "", "", "", "",
                 "", "", "", "", "", "", ""],
            ]
        )
        # One checkbox per row, all checked, values = session indices.
        self.assertContains(response, 'name="selected"', count=2)
        self.assertContains(
            response,
            'value="0" checked data-gender="female" data-age="over15" data-types="member"',
        )
        self.assertContains(
            response,
            'value="1" checked data-gender="male" data-age="under15" data-types=""',
        )

    def test_preview_metadata_from_clubhub_roles(self):
        content = _build_clubhub_xlsx([[
            "0007", "Kid", "Small", "20041003****", "Kvinna", "Gatan 1",
            "80321", "Gävle", "kid@example.com", "0737461137", "",
            "", "Tränare, Admin", "", "", "Small Kid",
        ]])
        response = self.client.post(
            reverse("people:import_preview"),
            {"file": SimpleUploadedFile("register.xlsx", content)},
        )
        self.assertContains(
            response,
            'data-gender="female" data-age="over15" data-types="trainer admin parent"',
        )

    def test_confirm_imports_only_selected_rows(self):
        self.upload_sportadmin(
            [
                ["20041003-****", "Kvinna", "Adult", "Person", "", "", "", "", "",
                 "", "", "", "adult@example.com", "", "", "", "", "", "", "", "",
                 "", "", "", "", "", "2025-04-10", ""],
                ["20180101-****", "Man", "Kid", "Person", "", "", "", "", "",
                 "", "", "", "", "", "", "", "", "", "", "", "",
                 "", "", "", "", "", "", ""],
            ]
        )
        response = self.confirm(selected=["1"])
        self.assertNotContains(response, "adult@example.com")
        self.assertTrue(Person.objects.filter(email="").filter(first_name="Kid").exists())
        self.assertFalse(Person.objects.filter(first_name="Adult").exists())

    def test_confirm_without_selection_imports_nothing(self):
        self.upload_sportadmin(
            [
                ["20041003-****", "Kvinna", "Adult", "Person", "", "", "", "", "",
                 "", "", "", "adult@example.com", "", "", "", "", "", "", "", "",
                 "", "", "", "", "", "2025-04-10", ""],
            ]
        )
        response = self.confirm()
        self.assertContains(response, "Inga personer valdes för import.")
        self.assertEqual(Person.objects.count(), 1)  # only the admin herself
